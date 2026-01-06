"""
Background extraction worker: Extract text from uploaded files.

Responsibility:
1. Monitor extraction_queue for extraction tasks
2. Read file from disk
3. Extract text using appropriate extractor (PDF, DOCX, TXT, etc.)
4. Create IngestionSource row with source_type=pdf_text and source_ref=file:{file_id}
5. If all files for a job are extracted, set job.status to READY_TO_INGEST
6. Continue processing until interrupted

This worker runs in a background thread, separate from the main FastAPI server
and the orchestration runner.
"""

import logging
from sqlalchemy.orm import Session

from app.storage.db import engine
from app.storage.models import Job, File, IngestionSource
from app.core.queues import extraction_queue, job_queue
from app.ingestion.document_format_extractors import DocumentExtractor

logger = logging.getLogger(__name__)


def start_extraction_worker():
    """
    Start the background extraction worker.
    
    This function runs in an infinite loop, processing extraction tasks
    from the extraction_queue. Should be started in a separate thread.
    """
    logger.info("Extraction worker started")
    
    while True:
        try:
            task = extraction_queue.get(timeout=5)
            process_extraction_task(task)
        except Exception as e:
            # Queue timeout (no tasks) is normal
            if "Empty" not in str(type(e).__name__):
                logger.error(f"Extraction worker error: {e}", exc_info=True)


def process_extraction_task(task: dict) -> None:
    """
    Process a single extraction task.
    
    Args:
        task: Dict with {job_id, file_id, task_type}
    """
    job_id = task.get("job_id")
    file_id = task.get("file_id")
    task_type = task.get("task_type", "extract_text")
    
    if not job_id or not file_id:
        logger.error(f"Invalid extraction task: {task}")
        return
    
    logger.info(f"Processing extraction task: job {job_id}, file {file_id}, type {task_type}")
    
    with Session(engine) as session:
        # Fetch the file
        file_row = session.query(File).filter(File.id == file_id).first()
        if not file_row:
            logger.error(f"File {file_id} not found")
            return
        
        if file_row.job_id != job_id:
            logger.error(f"File {file_id} does not belong to job {job_id}")
            return
        
        # Extract text from file
        try:
            extracted_text = extract_text_from_file(file_row.stored_path, file_row.file_type)
            
            if not extracted_text or not extracted_text.strip():
                logger.warning(f"Extracted no text from file {file_id}")
                return
            
            # Create IngestionSource row
            ingestion_source = IngestionSource(
                job_id=job_id,
                source_type="pdf_text",
                source_ref=f"file:{file_id}",
                raw_text=extracted_text,
                processed=False
            )
            session.add(ingestion_source)
            session.flush()
            
            logger.info(f"Created IngestionSource {ingestion_source.id} from file {file_id}")
            
            # Check if all files for this job are now extracted
            # (i.e., all files have corresponding IngestionSource rows)
            job_files = session.query(File).filter(File.job_id == job_id).all()
            job_ingestion_sources = session.query(IngestionSource).filter(
                IngestionSource.job_id == job_id,
                IngestionSource.source_type == "pdf_text"
            ).all()
            
            # Count extracted files (those with IngestionSource)
            extracted_file_ids = {
                int(s.source_ref.split(":")[1]) for s in job_ingestion_sources
                if s.source_ref.startswith("file:")
            }
            
            if len(extracted_file_ids) >= len(job_files):
                # All files extracted
                job = session.query(Job).filter(Job.id == job_id).first()
                if job and job.status != "READY_TO_INGEST":
                    job.status = "READY_TO_INGEST"
                    session.commit()
                    job_queue.put(job_id)  # Notify runner
                    logger.info(f"Job {job_id} is now READY_TO_INGEST; enqueued for processing")
                else:
                    session.commit()
            else:
                session.commit()
        
        except Exception as e:
            logger.error(f"Failed to extract text from file {file_id}: {e}", exc_info=True)
            # Mark job as failed? Or just skip this file?
            # For now, just log and continue
            session.commit()


def extract_text_from_file(file_path: str, file_type: str) -> str:
    """
    Extract text from a file based on its type.
    
    Args:
        file_path: Absolute path to the file
        file_type: File extension (pdf, txt, docx, etc.)
    
    Returns:
        Extracted text as a single string
    
    Raises:
        ValueError: If file cannot be extracted
    """
    file_type = file_type.lower().strip()
    
    if file_type == "pdf":
        # Extract from PDF
        pages = DocumentExtractor.extract_pdf(file_path)
        # Combine all pages with separators
        return "\n---PAGE BREAK---\n".join([content for content, _ in pages])
    
    elif file_type in ("docx", "doc"):
        # Extract from DOCX
        paragraphs = DocumentExtractor.extract_docx(file_path)
        # Combine all paragraphs
        return "\n".join([content for content, _ in paragraphs])
    
    elif file_type in ("txt", "text"):
        # Read plain text file
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    
    elif file_type in ("xlsx", "xls"):
        # Extract from spreadsheet
        try:
            import openpyxl
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError("openpyxl is required for Excel files")
        
        workbook = load_workbook(file_path)
        text_parts = []
        
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            text_parts.append(f"\n=== Sheet: {sheet} ===\n")
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell else "" for cell in row])
                if row_text.strip():
                    text_parts.append(row_text)
        
        return "\n".join(text_parts)
    
    else:
        raise ValueError(f"Unsupported file type: {file_type}")
