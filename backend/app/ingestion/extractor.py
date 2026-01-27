"""
Background extraction worker: Extract text from uploaded files.

Responsibility:
1. Monitor extraction_queue for extraction tasks
2. Read file from disk
3. Extract text using appropriate extractor (PDF, DOCX, TXT, etc.)
4. Create IngestionSource row with source_type=pdf_text and source_ref=file:{file_id}
5. If all files for a job are extracted, set job.status to READY_TO_INGEST
6. Continue processing until interrupted

This worker runs in a background thread, separate from the main FastAPI server
and the orchestration runner.
"""


from app.ingestion.document_format_extractors import DocumentExtractor


def extract_text_from_file(file_path: str, file_type: str) -> str:
    """
    Extract text from a file based on its type.
    
    Args:
        file_path: Absolute path to the file
        file_type: File extension (pdf, txt, docx, etc.)
    
    Returns:
        Extracted text as a single string
    
    Raises:
        ValueError: If file cannot be extracted
    """
    file_type = file_type.lower().strip()
    
    if file_type == "pdf":
        # Extract from PDF
        pages = DocumentExtractor.extract_pdf(file_path)
        # Combine all pages with separators
        return "\n---PAGE BREAK---\n".join([content for content, _ in pages])
    
    elif file_type in ("docx", "doc"):
        # Extract from DOCX
        paragraphs = DocumentExtractor.extract_docx(file_path)
        # Combine all paragraphs
        return "\n".join([content for content, _ in paragraphs])
    
    elif file_type in ("txt", "text"):
        # Read plain text file
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    
    elif file_type in ("xlsx", "xls"):
        # Extract from spreadsheet
        try:
            import openpyxl
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError("openpyxl is required for Excel files")
        
        workbook = load_workbook(file_path)
        text_parts = []
        
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            text_parts.append(f"\n=== Sheet: {sheet} ===\n")
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell else "" for cell in row])
                if row_text.strip():
                    text_parts.append(row_text)
        
        return "\n".join(text_parts)
    
    else:
        raise ValueError(f"Unsupported file type: {file_type}")
